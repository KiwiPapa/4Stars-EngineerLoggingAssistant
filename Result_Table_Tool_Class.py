# coding=utf-8
import os
import sys
import xlutils

import numpy as np
import openpyxl
import pandas as pd
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import load_workbook
from PyQt5 import QtCore, QtGui, QtWidgets, QtNetwork
from PyQt5.QtCore import QBasicTimer, QDateTime, Qt, QTimer
from PyQt5.QtPrintSupport import QPageSetupDialog, QPrintDialog, QPrinter
from PyQt5.QtWidgets import (QApplication, QColorDialog, QDialog, QFileDialog,
                             QFontDialog, QLabel, QLineEdit, QMainWindow,
                             QMessageBox, QPushButton, QRadioButton,
                             QTableWidgetItem, QTextEdit, QWidget)
from WELL_INTEGRITY_UI import Ui_MainWindow

class Main_window(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(Main_window, self).__init__()
        self.setupUi(self)
        self.setWindowOpacity(0.98)
        self.main_initialization()

    def main_initialization(self):
        self.radioButton.toggled.connect(lambda: self.btnstate_table(self.radioButton))
        self.radioButton_2.toggled.connect(lambda: self.btnstate_table(self.radioButton_2))

        self.pushButton_9.clicked.connect(self.open_file1)
        self.pushButton_10.clicked.connect(self.open_file2)
        self.pushButton_54.clicked.connect(self.open_file3)
        self.pushButton_55.clicked.connect(self.open_file4)
        self.pushButton_59.clicked.connect(self.open_file5)
        self.pushButton_60.clicked.connect(self.open_file6)

        self.pushButton_29.clicked.connect(self.reset_table_process)
        self.pushButton_52.clicked.connect(self.open_table_process_directory)  # 打开合并统计工区文件夹
        self.pushButton_56.clicked.connect(self.open_table_fusion_directory)  # 打开综合评价工区文件夹

        # 规范化2
        self.pushButton_57.clicked.connect(self.table_process3)
        self.pushButton_58.clicked.connect(self.table_fusion_reaction)

        self.action.triggered.connect(self.menubar_simple_instruction)
        self.action_2.triggered.connect(self.menubar_author_info)

        self.pushButton_61.clicked.connect(self.search_for_statistic_result)

    # 函数定义集结地
    def mkdir(self, path):
        path = path.strip()  # 去除首位空格
        path = path.rstrip("\\")  # 去除尾部 \ 符号
        isExists = os.path.exists(path)
        if not isExists:
            os.makedirs(path)
            # print(path + ' 创建成功')
            return True
        else:
            # print(path + ' 目录已存在')
            return False

    # 定义一个函数，增加重新计算后的厚度列
    def get_thickness(self, x):
        thickness = x['井段End'] - x['井段Start']
        return thickness

    # 一界面某井段评价函数
    def layer_evaluation1(self, df, start, end):
        df1 = df
        formation_Start = start
        formation_End = end
        # # 越界警告，不用加了，后面有考虑到越界情况
        # if df1.iloc[-1, df1.columns.get_loc('井段End')] < formation_End or df1.iloc[0, df1.columns.get_loc('井段Start')] > formation_Start:
        #     QMessageBox.information(self, '注意', '储层边界超过了单层评价表的范围。')
        # 截取我们想要的目标数据体
        df_temp = df1.loc[(df1['井段Start'] >= formation_Start) & (df1['井段Start'] <= formation_End), :]
        # 获取起始深度到第一层井段底界的结论
        df_temp_start_to_first_layer = df1.loc[(df1['井段Start'] <= formation_Start), :]
        if len(df_temp_start_to_first_layer) != 0:  # 若为空dataframe
            start_to_upper_result = df_temp_start_to_first_layer.loc[len(df_temp_start_to_first_layer), '结论']
        elif len(df_temp_start_to_first_layer) == 0: # 若不为空dataframe
            start_to_upper_result = df1.loc[1, '结论']
        # 获取calculation_Start所在段的声幅值
        df_temp_formation_Start = df1.loc[(df1['井段Start'] <= formation_Start) & (
                df1['井段End'] >= formation_Start), :]
        df_temp_formation_Start.reset_index(drop=True, inplace=True)  # 重新设置列索引#防止若截取中段，index不从0开始的bug
        # 补充储层界到井段的深度
        x, y = df_temp.shape
        df_temp = df_temp.reset_index()
        df_temp.drop(['index'], axis=1, inplace=True)
        if x != 0:  # 防止df_temp为空时，loc报错的bug
            first_layer_start = df_temp.loc[0, '井段Start']
        if x > 0 and first_layer_start != formation_Start:
            upper = pd.DataFrame({'井段(m)': ''.join([str(formation_Start), '-', str(first_layer_start)]),
                                  '厚度(m)': first_layer_start - formation_Start,
                                  '最大声幅(%)': df_temp_formation_Start.loc[0, '最大声幅(%)'],
                                  '最小声幅(%)': df_temp_formation_Start.loc[0, '最小声幅(%)'],
                                  '平均声幅(%)': df_temp_formation_Start.loc[0, '平均声幅(%)'],
                                  '结论': start_to_upper_result,
                                  '井段Start': formation_Start,
                                  '井段End': first_layer_start},
                                 index=[1])  # 自定义索引为：1 ，这里也可以不设置index
            df_temp.loc[len(df_temp) - 1, '井段End'] = formation_End
            df_temp = pd.concat([upper, df_temp], ignore_index=True)
            # df_temp = df_temp.append(new, ignore_index=True)  # ignore_index=True,表示不按原来的索引，从0开始自动递增
            df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
            # 修改df_temp的最末一行
            df_temp.loc[len(df_temp) - 1, '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp) - 1, '井段Start']), \
                                                              '-', str(df_temp.loc[len(df_temp) - 1, '井段End'])])
            df_temp.loc[len(df_temp) - 1, '厚度(m)'] = df_temp.loc[len(df_temp) - 1, '重计算厚度']
        elif x > 0 and first_layer_start == formation_Start:
            df_temp.loc[len(df_temp) - 1, '井段End'] = formation_End
            df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
            # 修改df_temp的最末一行
            df_temp.loc[len(df_temp) - 1, '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp) - 1, '井段Start']), \
                                                              '-', str(df_temp.loc[len(df_temp) - 1, '井段End'])])
            df_temp.loc[len(df_temp) - 1, '厚度(m)'] = df_temp.loc[len(df_temp) - 1, '重计算厚度']
        else:  # 储层包含在一个井段内的情况
            df_temp = pd.DataFrame({'井段(m)': ''.join([str(formation_Start), '-', str(formation_End)]),
                                    '厚度(m)': formation_End - formation_Start,
                                    '最大声幅(%)': df_temp_formation_Start.loc[0, '最大声幅(%)'],
                                    '最小声幅(%)': df_temp_formation_Start.loc[0, '最小声幅(%)'],
                                    '平均声幅(%)': df_temp_formation_Start.loc[0, '平均声幅(%)'],
                                    '结论': start_to_upper_result,
                                    '井段Start': formation_Start,
                                    '井段End': formation_End},
                                   index=[1])  # 自定义索引为：1 ，这里也可以不设置index
            df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
            # 修改df_temp的最末一行
            df_temp.loc[len(df_temp), '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp), '井段Start']),
                                                          '-', str(df_temp.loc[len(df_temp), '井段End'])])
            df_temp.loc[len(df_temp), '厚度(m)'] = df_temp.loc[len(df_temp), '重计算厚度']
        # print(df_temp)
        ratio_Series = df_temp.groupby(by=['结论'])['重计算厚度'].sum() / df_temp['重计算厚度'].sum() * 100

        if ratio_Series.__len__() == 2:
            if '好' not in ratio_Series:
                ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
            elif '中' not in ratio_Series:
                ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
            elif '差' not in ratio_Series:
                ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
        elif ratio_Series.__len__() == 1:
            if ('好' not in ratio_Series) & ('中' not in ratio_Series):
                ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
            elif ('好' not in ratio_Series) & ('差' not in ratio_Series):
                ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
            elif ('中' not in ratio_Series) & ('差' not in ratio_Series):
                ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                ratio_Series = ratio_Series.append(pd.Series({'差': 0}))

        # 条件判断，参数需要研究
        if ratio_Series['好'] >= 95:
            evaluation_of_formation = '好'
        elif ratio_Series['中'] >= 95:
            evaluation_of_formation = '中'
        elif ratio_Series['差'] >= 95:
            evaluation_of_formation = '差'
        elif (95 >= ratio_Series['好'] >= 5) & (95 >= ratio_Series['中'] >= 5) & (5 >= ratio_Series['差']):
            if ratio_Series['好'] >= ratio_Series['中']:
                evaluation_of_formation = '中到好，以好为主'
            elif ratio_Series['好'] <= ratio_Series['中']:
                evaluation_of_formation = '中到好，以中等为主'
        elif (95 >= ratio_Series['差'] >= 5) & (95 >= ratio_Series['中'] >= 5) & (5 >= ratio_Series['好']):
            if ratio_Series['差'] >= ratio_Series['中']:
                evaluation_of_formation = '中到差，以差为主'
            elif ratio_Series['差'] <= ratio_Series['中']:
                evaluation_of_formation = '中到差，以中等为主'
        elif (95 >= ratio_Series['好'] >= 5) & (95 >= ratio_Series['差'] >= 5) & (5 >= ratio_Series['中']):
            if ratio_Series['好'] >= ratio_Series['差']:
                evaluation_of_formation = '好到差，以好为主'
            elif ratio_Series['好'] <= ratio_Series['差']:
                evaluation_of_formation = '好到差，以差为主'
        elif (95 > ratio_Series['好'] > 5) & (95 > ratio_Series['差'] > 5) & (95 > ratio_Series['中'] > 5):
            evaluation_of_formation = '好到中到差'
        elif (95 > ratio_Series['好'] > 5) & (5 >= ratio_Series['差']) & (5 >= ratio_Series['中']):
            evaluation_of_formation = '好到中到差，以好为主'
        elif (5 >= ratio_Series['好']) & (5 >= ratio_Series['差']) & (95 > ratio_Series['中'] > 5):
            evaluation_of_formation = '好到中到差，以中等为主'
        elif (5 >= ratio_Series['好']) & (95 > ratio_Series['差'] > 5) & (5 >= ratio_Series['中']):
            evaluation_of_formation = '好到中到差，以差为主'
        return ratio_Series, evaluation_of_formation

    # 二界面某井段评价函数
    def layer_evaluation2(self, df, start, end):
        df1 = df
        formation_Start = start
        formation_End = end
        df_temp = df1.loc[(df1['井段Start'] >= formation_Start) & (df1['井段Start'] <= formation_End), :]
        # 获取起始深度到第一层井段底界的结论
        df_temp_start_to_first_layer = df1.loc[(df1['井段Start'] <= formation_Start), :]
        if len(df_temp_start_to_first_layer) != 0:  # 若为空dataframe
            start_to_upper_result = df_temp_start_to_first_layer.loc[len(df_temp_start_to_first_layer), '结论']
        elif len(df_temp_start_to_first_layer) == 0:  # 若不为空dataframe
            start_to_upper_result = df1.loc[1, '结论']
        # 获取calculation_Start所在段的声幅值
        df_temp_formation_Start = df1.loc[(df1['井段Start'] <= formation_Start) & (
                df1['井段End'] >= formation_Start), :]
        df_temp_formation_Start.reset_index(drop=True, inplace=True)  # 重新设置列索引#防止若截取中段，index不从0开始的bug
        # 补充储层界到井段的深度
        x, y = df_temp.shape
        df_temp = df_temp.reset_index()
        df_temp.drop(['index'], axis=1, inplace=True)
        if x != 0:  # 防止df_temp为空时，loc报错的bug
            first_layer_start = df_temp.loc[0, '井段Start']
        if x > 0 and first_layer_start != formation_Start:
            upper = pd.DataFrame({'井段(m)': ''.join([str(formation_Start), '-', str(first_layer_start)]),
                                  '厚度(m)': first_layer_start - formation_Start,
                                  '最大指数': df_temp_formation_Start.loc[0, '最大指数'],
                                  '最小指数': df_temp_formation_Start.loc[0, '最小指数'],
                                  '平均指数': df_temp_formation_Start.loc[0, '平均指数'],
                                  '结论': start_to_upper_result,
                                  '井段Start': formation_Start,
                                  '井段End': first_layer_start},
                                 index=[1])  # 自定义索引为：1 ，这里也可以不设置index
            df_temp.loc[len(df_temp) - 1, '井段End'] = formation_End
            df_temp = pd.concat([upper, df_temp], ignore_index=True)
            # df_temp = df_temp.append(new, ignore_index=True)  # ignore_index=True,表示不按原来的索引，从0开始自动递增
            df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
            # 修改df_temp的最末一行
            df_temp.loc[len(df_temp) - 1, '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp) - 1, '井段Start']), \
                                                              '-', str(df_temp.loc[len(df_temp) - 1, '井段End'])])
            df_temp.loc[len(df_temp) - 1, '厚度(m)'] = df_temp.loc[len(df_temp) - 1, '重计算厚度']
        elif x > 0 and first_layer_start == formation_Start:
            df_temp.loc[len(df_temp) - 1, '井段End'] = formation_End
            df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
            # 修改df_temp的最末一行
            df_temp.loc[len(df_temp) - 1, '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp) - 1, '井段Start']), \
                                                              '-', str(df_temp.loc[len(df_temp) - 1, '井段End'])])
            df_temp.loc[len(df_temp) - 1, '厚度(m)'] = df_temp.loc[len(df_temp) - 1, '重计算厚度']
        else:  # 储层包含在一个井段内的情况
            df_temp = pd.DataFrame({'井段(m)': ''.join([str(formation_Start), '-', str(formation_End)]),
                                    '厚度(m)': formation_End - formation_Start,
                                    '最大指数': df_temp_formation_Start.loc[0, '最大指数'],
                                    '最小指数': df_temp_formation_Start.loc[0, '最小指数'],
                                    '平均指数': df_temp_formation_Start.loc[0, '平均指数'],
                                    '结论': start_to_upper_result,
                                    '井段Start': formation_Start,
                                    '井段End': formation_End},
                                   index=[1])  # 自定义索引为：1 ，这里也可以不设置index
            df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
            # 修改df_temp的最末一行
            df_temp.loc[len(df_temp), '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp), '井段Start']),
                                                          '-', str(df_temp.loc[len(df_temp), '井段End'])])
            df_temp.loc[len(df_temp), '厚度(m)'] = df_temp.loc[len(df_temp), '重计算厚度']
        # print(df_temp)
        ratio_Series = df_temp.groupby(by=['结论'])['重计算厚度'].sum() / df_temp['重计算厚度'].sum() * 100
        if '不确定' not in ratio_Series:
            if ratio_Series.__len__() == 2:
                if '好' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                elif '中' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                elif '差' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
            elif ratio_Series.__len__() == 1:
                if ('好' not in ratio_Series) & ('中' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                elif ('好' not in ratio_Series) & ('差' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
                elif ('中' not in ratio_Series) & ('差' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
        else:
            if ratio_Series.__len__() == 1:
                ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
            elif ratio_Series.__len__() == 2:
                if ('好' not in ratio_Series) & ('中' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                elif ('好' not in ratio_Series) & ('差' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
                elif ('中' not in ratio_Series) & ('差' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
            elif ratio_Series.__len__() == 3:
                if '好' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                elif '中' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                elif '差' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))

        # 条件判断，参数需要研究
        if ratio_Series['好'] >= 95:
            evaluation_of_formation = '好'
        elif ratio_Series['中'] >= 95:
            evaluation_of_formation = '中'
        elif ratio_Series['差'] >= 95:
            evaluation_of_formation = '差'
        elif (95 >= ratio_Series['好'] >= 5) & (95 >= ratio_Series['中'] >= 5) & (5 >= ratio_Series['差']):
            if ratio_Series['好'] >= ratio_Series['中']:
                evaluation_of_formation = '中到好，以好为主'
            elif ratio_Series['好'] <= ratio_Series['中']:
                evaluation_of_formation = '中到好，以中等为主'
        elif (95 >= ratio_Series['差'] >= 5) & (95 >= ratio_Series['中'] >= 5) & (5 >= ratio_Series['好']):
            if ratio_Series['差'] >= ratio_Series['中']:
                evaluation_of_formation = '中到差，以差为主'
            elif ratio_Series['差'] <= ratio_Series['中']:
                evaluation_of_formation = '中到差，以中等为主'
        elif (95 >= ratio_Series['好'] >= 5) & (95 >= ratio_Series['差'] >= 5) & (5 >= ratio_Series['中']):
            if ratio_Series['好'] >= ratio_Series['差']:
                evaluation_of_formation = '好到差，以好为主'
            elif ratio_Series['好'] <= ratio_Series['差']:
                evaluation_of_formation = '好到差，以差为主'
        elif (95 > ratio_Series['好'] > 5) & (95 > ratio_Series['差'] > 5) & (95 > ratio_Series['中'] > 5):
            evaluation_of_formation = '好到中到差'
        elif (95 > ratio_Series['好'] > 5) & (5 >= ratio_Series['差']) & (5 >= ratio_Series['中']):
            evaluation_of_formation = '好到中到差，以好为主'
        elif (5 >= ratio_Series['好']) & (5 >= ratio_Series['差']) & (95 > ratio_Series['中'] > 5):
            evaluation_of_formation = '好到中到差，以中等为主'
        elif (5 >= ratio_Series['好']) & (95 > ratio_Series['差'] > 5) & (5 >= ratio_Series['中']):
            evaluation_of_formation = '好到中到差，以差为主'
        elif ratio_Series['不确定'] >= 1:
            evaluation_of_formation = '不确定'
        return ratio_Series, evaluation_of_formation

    # 关于
    def menubar_simple_instruction(self):
        QMessageBox.information(self, "简单介绍", "主要实现了固井质量评价的成果表综合统计、拼接和分段统计、查询优中差比例等功能")

    def menubar_author_info(self):
        QMessageBox.information(self, "联系方式", "软件开发: 杨艺 \n电话：18580367621，邮箱：978030836@qq.com\n软件测试: 刘恒 王参文 何强")

    # 打开表格函数集合
    ##############################
    # 单层评价表拼接和分段统计界面
    def open_file1(self):
        fnames = QFileDialog.getOpenFileNames(self, '打开第一个文件', './')  # 注意这里返回值是元组
        if fnames[0]:
            for fname in fnames[0]:
                self.lineEdit_48.setText(fname)

    def open_file2(self):
        fnames = QFileDialog.getOpenFileNames(self, '打开第二个文件', './')  # 注意这里返回值是元组
        if fnames[0]:
            for fname in fnames[0]:
                self.lineEdit_52.setText(fname)

    # 固井质量综合评价界面
    def open_file3(self):
        fnames = QFileDialog.getOpenFileNames(self, '打开第一个文件', './')  # 注意这里返回值是元组
        if fnames[0]:
            for fname in fnames[0]:
                self.lineEdit_75.setText(fname)

    def open_file4(self):
        fnames = QFileDialog.getOpenFileNames(self, '打开第二个文件', './')  # 注意这里返回值是元组
        if fnames[0]:
            for fname in fnames[0]:
                self.lineEdit_73.setText(fname)

    # 查询优中差比例界面
    def open_file5(self):
        fnames = QFileDialog.getOpenFileNames(self, '打开第一个文件', './')  # 注意这里返回值是元组
        if fnames[0]:
            for fname in fnames[0]:
                self.lineEdit_77.setText(fname)

    def open_file6(self):
        fnames = QFileDialog.getOpenFileNames(self, '打开第二个文件', './')  # 注意这里返回值是元组
        if fnames[0]:
            for fname in fnames[0]:
                self.lineEdit_76.setText(fname)
    ##############################

    # 规范化1
    ##############################
    def xls_formatting_first_layer(self, path):  # 用于一界面表格表头文字规范
        # 打开想要更改的excel文件
        old_excel = xlrd.open_workbook(path, formatting_info=True)
        # 将操作文件对象拷贝，变成可写的workbook对象
        new_excel = copy(old_excel)
        # 获得第一个sheet的对象
        ws = new_excel.get_sheet(0)
        # 写入数据
        ws.write(2, 0, '解释序号')
        ws.write(2, 1, '井段(m)')
        ws.write(2, 2, '厚度(m)')
        ws.write(2, 3, '最大声幅(%)')
        ws.write(2, 4, '最小声幅(%)')
        ws.write(2, 5, '平均声幅(%)')
        ws.write(2, 6, '结论')
        # 另存为excel文件，并将文件命名
        new_excel.save(path)

    def xls_formatting_second_layer(self, path):  # 用于二界面表格表头文字规范
        # 打开想要更改的excel文件
        old_excel = xlrd.open_workbook(path, formatting_info=True)
        # 将操作文件对象拷贝，变成可写的workbook对象
        new_excel = copy(old_excel)
        # 获得第一个sheet的对象
        ws = new_excel.get_sheet(0)
        # 写入数据
        ws.write(2, 0, '解释序号')
        ws.write(2, 1, '井段(m)')
        ws.write(2, 2, '厚度(m)')
        ws.write(2, 3, '最大指数')
        ws.write(2, 4, '最小指数')
        ws.write(2, 5, '平均指数')
        ws.write(2, 6, '结论')
        # 另存为excel文件，并将文件命名
        new_excel.save(path)

    def btnstate_table(self, btn):
        # 输出按钮1与按钮2的状态，选中还是没选中
        if btn.text() == '一界面':
            if btn.isChecked() == True:
                print(btn.text() + " 被选中")
                self.pushButton_14.clicked.connect(self.calculate_for_first_layer)
                self.pushButton_27.clicked.connect(self.table_process1)
            else:
                pass

        if btn.text() == "二界面":
            if btn.isChecked() == True:
                print(btn.text() + " 被选中")
                self.pushButton_14.clicked.connect(self.calculate_for_second_layer)
                self.pushButton_27.clicked.connect(self.table_process2)
            else:
                pass

    def reset_table_process(self):
        try:
            self.radioButton.toggled.disconnect(lambda: self.btnstate_table(self.radioButton))
        except:
            print('Error1')
        else:
            print('disconnected1')

        try:
            self.radioButton_2.toggled.disconnect(lambda: self.btnstate_table(self.radioButton_2))
        except:
            print('Error2')
        else:
            print('disconnected2')

        try:
            self.pushButton_14.clicked.disconnect(self.calculate_for_first_layer)
        except:
            print('Error3')
        else:
            print('disconnected3')

        try:
            self.pushButton_27.clicked.disconnect(self.table_process1)
        except:
            print('Error4')
        else:
            print('disconnected4')

        try:
            self.pushButton_14.clicked.disconnect(self.calculate_for_second_layer)
        except:
            print('Error5')
        else:
            print('disconnected5')

        try:
            self.pushButton_27.clicked.disconnect(self.table_process2)
        except:
            print('Error6')
        else:
            print('disconnected6')
        self.radioButton.toggled.connect(lambda: self.btnstate_table(self.radioButton))
        self.radioButton_2.toggled.connect(lambda: self.btnstate_table(self.radioButton_2))

    def table_process1(self):
        fileDir1 = self.lineEdit_48.text()
        fileDir2 = self.lineEdit_52.text()
        self.xls_formatting_first_layer(fileDir1)
        if fileDir1 != fileDir2:
            self.xls_formatting_first_layer(fileDir2)
        QMessageBox.information(self, "提示", "一界面表格数据规范化完毕")

    def table_process2(self):
        fileDir1 = self.lineEdit_48.text()
        fileDir2 = self.lineEdit_52.text()
        self.xls_formatting_second_layer(fileDir1)
        if fileDir1 != fileDir2:
            self.xls_formatting_second_layer(fileDir2)
        QMessageBox.information(self, "提示", "二界面表格数据规范化完毕")
    ##############################

    # 规范化2
    ##############################
    def table_process3(self):
        if self.lineEdit_75.text() != '' and self.lineEdit_73.text() == '':
            fileDir1 = self.lineEdit_75.text()
            self.xls_formatting_first_layer(fileDir1)
            QMessageBox.information(self, "提示", "一界面表格数据规范化完毕")
        elif self.lineEdit_73.text() != '' and self.lineEdit_75.text() == '':
            fileDir2 = self.lineEdit_73.text()
            self.xls_formatting_second_layer(fileDir2)
            QMessageBox.information(self, "提示", "二界面表格数据规范化完毕")
        elif self.lineEdit_73.text() != '' and self.lineEdit_75.text() != '':
            fileDir1 = self.lineEdit_75.text()
            self.xls_formatting_first_layer(fileDir1)
            fileDir2 = self.lineEdit_73.text()
            self.xls_formatting_second_layer(fileDir2)
            QMessageBox.information(self, "提示", "一二界面表格数据都规范化完毕")

    def calculate_for_first_layer(self):
        splicing_Depth = float(self.lineEdit_45.text())

        fileDir1 = self.lineEdit_48.text()
        fileDir2 = self.lineEdit_52.text()

        df1 = pd.read_excel(fileDir1, header=2)
        df1.drop([0], inplace=True)
        df1 = df1.dropna(axis=0, how='any')  # 删除dataframe里NaN的所有行
        df1.loc[:, '井段(m)'] = df1['井段(m)'].str.replace(' ', '')  # 消除数据中空格
        # if len(df1) % 2 == 0:#如果len(df1)为偶数需要删除最后一行NaN，一行的情况不用删
        #     df1.drop([len(df1)], inplace=True)
        df1['井段Start'] = df1['井段(m)'].map(lambda x: x.split("-")[0])
        df1['井段End'] = df1['井段(m)'].map(lambda x: x.split("-")[1])
        # 表格数据清洗
        df1.loc[:, "井段Start"] = df1["井段Start"].str.replace(" ", "").astype('float')
        df1.loc[:, "井段End"] = df1["井段End"].str.replace(" ", "").astype('float')

        # 截取拼接点以上的数据体
        df_temp1 = df1.loc[(df1['井段Start'] <= splicing_Depth), :].copy()  # 加上copy()可防止直接修改df报错
        # print(df_temp1)

        #####################################################
        df2 = pd.read_excel(fileDir2, header=2)
        df2.drop([0], inplace=True)
        df2 = df2.dropna(axis=0, how='any')  # 删除dataframe里NaN的所有行
        df2.loc[:, '井段(m)'] = df2['井段(m)'].str.replace(' ', '')  # 消除数据中空格
        # if len(df2) % 2 == 0:#如果len(df2)为偶数需要删除最后一行NaN，一行的情况不用删
        #     df2.drop([len(df2)], inplace=True)
        df2['井段Start'] = df2['井段(m)'].map(lambda x: x.split("-")[0])
        df2['井段End'] = df2['井段(m)'].map(lambda x: x.split("-")[1])
        # 表格数据清洗
        df2.loc[:, "井段Start"] = df2["井段Start"].str.replace(" ", "").astype('float')
        df2.loc[:, "井段End"] = df2["井段End"].str.replace(" ", "").astype('float')

        # 截取拼接点以下的数据体
        df_temp2 = df2.loc[(df2['井段End'] >= splicing_Depth), :].copy()  # 加上copy()可防止直接修改df报错
        df_temp2.reset_index(drop=True, inplace=True)  # 重新设置列索引

        # print(df_temp2)

        df_all = df_temp1.append(df_temp2)
        df_all.reset_index(drop=True, inplace=True)  # 重新设置列索引
        # 对df_all进行操作
        df_all.loc[len(df_temp1) - 1, '井段(m)'] = ''.join([str(df_all.loc[len(df_temp1) - 1, '井段Start']), '-', \
                                                          str(df_all.loc[len(df_temp1), '井段End'])])
        df_all.loc[len(df_temp1) - 1, '厚度(m)'] = df_all.loc[len(df_temp1), '井段End'] - \
                                                 df_all.loc[len(df_temp1) - 1, '井段Start']
        df_all.loc[len(df_temp1) - 1, '最大声幅(%)'] = max(df_all.loc[len(df_temp1), '最大声幅(%)'], \
                                                       df_all.loc[len(df_temp1) - 1, '最大声幅(%)'])
        df_all.loc[len(df_temp1) - 1, '最小声幅(%)'] = min(df_all.loc[len(df_temp1), '最小声幅(%)'], \
                                                       df_all.loc[len(df_temp1) - 1, '最小声幅(%)'])
        df_all.loc[len(df_temp1) - 1, '平均声幅(%)'] = np.add(df_all.loc[len(df_temp1), '平均声幅(%)'], \
                                                          df_all.loc[len(df_temp1) - 1, '平均声幅(%)']) / 2
        df_all.loc[len(df_temp1) - 1, '井段End'] = df_all.loc[len(df_temp1), '井段End']  # 解决后续重计算厚度计算bug
        df_all.drop(len(df_temp1), inplace=True)
        df_all.set_index(["解释序号"], inplace=True)
        df_all.reset_index(drop=True, inplace=True)  # 重新设置列索引
        # print(df_all.columns)

        #################################################################
        # 在指定深度段统计

        # calculation_Start = float(input('请输入开始统计深度'))
        # calculation_End = float(input('请输入结束统计深度'))
        calculation_Start = float(self.lineEdit_46.text())
        calculation_End = float(self.lineEdit_47.text())

        start_Evaluation = df_all.loc[0, '井段(m)'].split('-')[0]
        end_Evaluation = df_all.loc[len(df_all) - 1, '井段(m)'].split('-')[1]
        if (calculation_End <= float(end_Evaluation)) & (calculation_Start >= float(start_Evaluation)):
            df_temp = df_all.loc[(df_all['井段Start'] >= calculation_Start) & (df_all['井段Start'] <= calculation_End), :]
            # 获取起始深度到第一层井段底界的结论
            df_temp_start_to_first_layer = df_all.loc[(df_all['井段Start'] <= calculation_Start), :]
            start_to_upper_result = df_temp_start_to_first_layer.loc[len(df_temp_start_to_first_layer) - 1, '结论']
            # 获取calculation_Start所在段的声幅值
            df_temp_calculation_Start = df_all.loc[(df_all['井段Start'] <= calculation_Start) & (
                    df_all['井段End'] >= calculation_Start), :]
            df_temp_calculation_Start.reset_index(drop=True, inplace=True)  # 重新设置列索引#防止若截取中段，index不从0开始的bug
            # 补充储层界到井段的深度
            x, y = df_temp.shape
            df_temp = df_temp.reset_index()
            df_temp.drop(['index'], axis=1, inplace=True)
            if x != 0:  # 防止df_temp为空时，loc报错的bug
                first_layer_start = df_temp.loc[0, '井段Start']
            if x > 1 and first_layer_start != calculation_Start:
                upper = pd.DataFrame({'井段(m)': ''.join([str(calculation_Start), '-', str(first_layer_start)]),
                                      '厚度(m)': first_layer_start - calculation_Start,
                                      '最大声幅(%)': df_temp_calculation_Start.loc[0, '最大声幅(%)'],
                                      '最小声幅(%)': df_temp_calculation_Start.loc[0, '最小声幅(%)'],
                                      '平均声幅(%)': df_temp_calculation_Start.loc[0, '平均声幅(%)'],
                                      '结论': start_to_upper_result,
                                      '井段Start': calculation_Start,
                                      '井段End': first_layer_start},
                                     index=[1])  # 自定义索引为：1 ，这里也可以不设置index
                df_temp.loc[len(df_temp) - 1, '井段End'] = calculation_End
                df_temp = pd.concat([upper, df_temp], ignore_index=True)
                # df_temp = df_temp.append(new, ignore_index=True)  # ignore_index=True,表示不按原来的索引，从0开始自动递增
                df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
                # 修改df_temp的最末一行
                df_temp.loc[len(df_temp) - 1, '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp) - 1, '井段Start']), \
                                                                  '-', str(df_temp.loc[len(df_temp) - 1, '井段End'])])
                df_temp.loc[len(df_temp) - 1, '厚度(m)'] = df_temp.loc[len(df_temp) - 1, '重计算厚度']
            elif x > 1 and first_layer_start == calculation_Start:
                df_temp.loc[len(df_temp) - 1, '井段End'] = calculation_End
                df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
                # 修改df_temp的最末一行
                df_temp.loc[len(df_temp) - 1, '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp) - 1, '井段Start']), \
                                                                  '-', str(df_temp.loc[len(df_temp) - 1, '井段End'])])
                df_temp.loc[len(df_temp) - 1, '厚度(m)'] = df_temp.loc[len(df_temp) - 1, '重计算厚度']
            else:  # 储层包含在一个井段内的情况
                df_temp = pd.DataFrame({'井段(m)': ''.join([str(calculation_Start), '-', str(calculation_End)]),
                                        '厚度(m)': calculation_End - calculation_Start,
                                        '最大声幅(%)': df_temp_calculation_Start.loc[0, '最大声幅(%)'],
                                        '最小声幅(%)': df_temp_calculation_Start.loc[0, '最小声幅(%)'],
                                        '平均声幅(%)': df_temp_calculation_Start.loc[0, '平均声幅(%)'],
                                        '结论': start_to_upper_result,
                                        '井段Start': calculation_Start,
                                        '井段End': calculation_End},
                                       index=[1])  # 自定义索引为：1 ，这里也可以不设置index
                df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
                # 修改df_temp的最末一行
                df_temp.loc[len(df_temp), '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp), '井段Start']),
                                                              '-', str(df_temp.loc[len(df_temp), '井段End'])])
                df_temp.loc[len(df_temp), '厚度(m)'] = df_temp.loc[len(df_temp), '重计算厚度']
            print(df_temp)
            ratio_Series = df_temp.groupby(by=['结论'])['重计算厚度'].sum() / df_temp['重计算厚度'].sum() * 100
            if ratio_Series.__len__() == 2:
                if '好' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                elif '中' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                elif '差' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
            elif ratio_Series.__len__() == 1:
                if ('好' not in ratio_Series) & ('中' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                elif ('好' not in ratio_Series) & ('差' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
                elif ('中' not in ratio_Series) & ('差' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
        # print(ratio_Series)

        # 统计结论
        actual_Hao = str(round((calculation_End - calculation_Start) * (ratio_Series['好'] / 100), 2))
        Hao_Ratio = str(round(ratio_Series['好'], 2))

        actual_Zhong = str(round((calculation_End - calculation_Start) * (ratio_Series['中'] / 100), 2))
        Zhong_Ratio = str(round(ratio_Series['中'], 2))

        actual_Cha = str(round(calculation_End - calculation_Start - float(actual_Hao) - float(actual_Zhong), 2))
        Cha_Ratio = str(round(100.00 - float(Hao_Ratio) - float(Zhong_Ratio), 2))

        PATH = '.\\resources\\模板\\'
        wb = openpyxl.load_workbook(PATH + '1统模板.xlsx')
        sheet = wb[wb.sheetnames[0]]
        sheet['A1'] = ''.join(['第一界面水泥胶结统计表（', str(calculation_Start), '-', str(calculation_End), 'm）'])
        sheet['C4'] = actual_Hao
        sheet['D4'] = Hao_Ratio
        sheet['C5'] = actual_Zhong
        sheet['D5'] = Zhong_Ratio
        sheet['C6'] = actual_Cha
        sheet['D6'] = Cha_Ratio

        self.mkdir('.\\WorkSpace\\合并统计工区')
        wb.save(
            '.\\WorkSpace\\合并统计工区\\一界面水泥胶结统计表(' + str(calculation_Start) + '-' + str(calculation_End) + 'm).xlsx')

        # 保存指定起始截止深度的单层统计表
        df_temp.drop(['井段Start', '井段End', '重计算厚度'], axis=1, inplace=True)
        df_temp.reset_index(drop=True, inplace=True)  # 重新设置列索引
        df_temp.index = df_temp.index + 1
        writer = pd.ExcelWriter(
            '.\\WorkSpace\\合并统计工区\\一界面水泥胶结单层评价表(' + str(calculation_Start) + '-' + str(calculation_End) + 'm).xlsx')
        df_temp.to_excel(writer, 'Sheet1')
        writer.save()

        QMessageBox.information(self, "提示", "运行完毕，请查看WorkSpace")

    def calculate_for_second_layer(self):
        splicing_Depth = float(self.lineEdit_45.text())

        fileDir1 = self.lineEdit_48.text()
        fileDir2 = self.lineEdit_52.text()

        df1 = pd.read_excel(fileDir1, header=2)
        df1.drop([0], inplace=True)
        if df1.loc[1, '结论'] == '不确定':
            df1.drop([1], inplace=True)
        df1 = df1.dropna(axis=0, how='any')  # 删除dataframe里NaN的所有行
        df1.loc[:, '井段(m)'] = df1['井段(m)'].str.replace(' ', '')  # 消除数据中空格
        # if len(df1) % 2 == 0:#如果len(df1)为偶数需要删除最后一行NaN，一行的情况不用删
        #     df1.drop([len(df1)], inplace=True)
        df1['井段Start'] = df1['井段(m)'].map(lambda x: x.split("-")[0])
        df1['井段End'] = df1['井段(m)'].map(lambda x: x.split("-")[1])
        # 表格数据清洗
        df1.loc[:, "井段Start"] = df1["井段Start"].str.replace(" ", "").astype('float')
        df1.loc[:, "井段End"] = df1["井段End"].str.replace(" ", "").astype('float')

        # 截取拼接点以上的数据体
        df_temp1 = df1.loc[(df1['井段Start'] <= splicing_Depth), :].copy()  # 加上copy()可防止直接修改df报错
        # print(df_temp1)

        #####################################################
        df2 = pd.read_excel(fileDir2, header=2)
        df2.drop([0], inplace=True)
        df2 = df2.dropna(axis=0, how='any')  # 删除dataframe里NaN的所有行
        df2.loc[:, '井段(m)'] = df2['井段(m)'].str.replace(' ', '')  # 消除数据中空格
        # if len(df2) % 2 == 0:#如果len(df2)为偶数需要删除最后一行NaN，一行的情况不用删
        #     df2.drop([len(df2)], inplace=True)
        df2['井段Start'] = df2['井段(m)'].map(lambda x: x.split("-")[0])
        df2['井段End'] = df2['井段(m)'].map(lambda x: x.split("-")[1])
        # 表格数据清洗
        df2.loc[:, "井段Start"] = df2["井段Start"].str.replace(" ", "").astype('float')
        df2.loc[:, "井段End"] = df2["井段End"].str.replace(" ", "").astype('float')

        # 截取拼接点以下的数据体
        df_temp2 = df2.loc[(df2['井段End'] >= splicing_Depth), :].copy()  # 加上copy()可防止直接修改df报错
        df_temp2.reset_index(drop=True, inplace=True)  # 重新设置列索引

        # print(df_temp2)

        df_all = df_temp1.append(df_temp2)
        df_all.reset_index(drop=True, inplace=True)  # 重新设置列索引
        # 对df_all进行操作
        df_all.loc[len(df_temp1) - 1, '井段(m)'] = ''.join([str(df_all.loc[len(df_temp1) - 1, '井段Start']), '-', \
                                                          str(df_all.loc[len(df_temp1), '井段End'])])
        df_all.loc[len(df_temp1) - 1, '厚度(m)'] = df_all.loc[len(df_temp1), '井段End'] - \
                                                 df_all.loc[len(df_temp1) - 1, '井段Start']
        df_all.loc[len(df_temp1) - 1, '最大指数'] = max(df_all.loc[len(df_temp1), '最大指数'], \
                                                    df_all.loc[len(df_temp1) - 1, '最大指数'])
        df_all.loc[len(df_temp1) - 1, '最小指数'] = min(df_all.loc[len(df_temp1), '最小指数'], \
                                                    df_all.loc[len(df_temp1) - 1, '最小指数'])
        df_all.loc[len(df_temp1) - 1, '平均指数'] = np.add(df_all.loc[len(df_temp1), '平均指数'], \
                                                       df_all.loc[len(df_temp1) - 1, '平均指数']) / 2

        df_all.drop(len(df_temp1), inplace=True)
        df_all.set_index(["解释序号"], inplace=True)
        df_all.reset_index(drop=True, inplace=True)  # 重新设置列索引
        # print(df_all.columns)

        #################################################################
        # 在指定深度段统计

        # calculation_Start = float(input('请输入开始统计深度'))
        # calculation_End = float(input('请输入结束统计深度'))
        calculation_Start = float(self.lineEdit_46.text())
        calculation_End = float(self.lineEdit_47.text())

        start_Evaluation = df_all.loc[0, '井段(m)'].split('-')[0]
        end_Evaluation = df_all.loc[len(df_all) - 1, '井段(m)'].split('-')[1]
        if (calculation_End <= float(end_Evaluation)) & (calculation_Start >= float(start_Evaluation)):
            df_temp = df_all.loc[(df_all['井段Start'] >= calculation_Start) & (df_all['井段Start'] <= calculation_End),
                      :]
            # 获取起始深度到第一层井段底界的结论
            df_temp_start_to_first_layer = df_all.loc[(df_all['井段Start'] <= calculation_Start), :]
            start_to_upper_result = df_temp_start_to_first_layer.loc[len(df_temp_start_to_first_layer) - 1, '结论']
            # 获取calculation_Start所在段的声幅值
            df_temp_calculation_Start = df_all.loc[(df_all['井段Start'] <= calculation_Start) & (
                    df_all['井段End'] >= calculation_Start), :]
            df_temp_calculation_Start.reset_index(drop=True, inplace=True)  # 重新设置列索引#防止若截取中段，index不从0开始的bug
            # 补充储层界到井段的深度
            x, y = df_temp.shape
            df_temp = df_temp.reset_index()
            df_temp.drop(['index'], axis=1, inplace=True)
            if x != 0:  # 防止df_temp为空时，loc报错的bug
                first_layer_start = df_temp.loc[0, '井段Start']
            if x > 1 and first_layer_start != calculation_Start:
                upper = pd.DataFrame({'井段(m)': ''.join([str(calculation_Start), '-', str(first_layer_start)]),
                                      '厚度(m)': first_layer_start - calculation_Start,
                                      '最大指数': df_temp_calculation_Start.loc[0, '最大指数'],
                                      '最小指数': df_temp_calculation_Start.loc[0, '最小指数'],
                                      '平均指数': df_temp_calculation_Start.loc[0, '平均指数'],
                                      '结论': start_to_upper_result,
                                      '井段Start': calculation_Start,
                                      '井段End': first_layer_start},
                                     index=[1])  # 自定义索引为：1 ，这里也可以不设置index
                df_temp.loc[len(df_temp) - 1, '井段End'] = calculation_End
                df_temp = pd.concat([upper, df_temp], ignore_index=True)
                # df_temp = df_temp.append(new, ignore_index=True)  # ignore_index=True,表示不按原来的索引，从0开始自动递增
                df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
                # 修改df_temp的最末一行
                df_temp.loc[len(df_temp) - 1, '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp) - 1, '井段Start']), \
                                                                  '-',
                                                                  str(df_temp.loc[len(df_temp) - 1, '井段End'])])
                df_temp.loc[len(df_temp) - 1, '厚度(m)'] = df_temp.loc[len(df_temp) - 1, '重计算厚度']
            elif x > 1 and first_layer_start == calculation_Start:
                df_temp.loc[len(df_temp) - 1, '井段End'] = calculation_End
                df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
                # 修改df_temp的最末一行
                df_temp.loc[len(df_temp) - 1, '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp) - 1, '井段Start']), \
                                                                  '-',
                                                                  str(df_temp.loc[len(df_temp) - 1, '井段End'])])
                df_temp.loc[len(df_temp) - 1, '厚度(m)'] = df_temp.loc[len(df_temp) - 1, '重计算厚度']
            else:  # 储层包含在一个井段内的情况
                df_temp = pd.DataFrame({'井段(m)': ''.join([str(calculation_Start), '-', str(calculation_End)]),
                                        '厚度(m)': calculation_End - calculation_Start,
                                        '最大指数': df_temp_calculation_Start.loc[0, '最大指数'],
                                        '最小指数': df_temp_calculation_Start.loc[0, '最小指数'],
                                        '平均指数': df_temp_calculation_Start.loc[0, '平均指数'],
                                        '结论': start_to_upper_result,
                                        '井段Start': calculation_Start,
                                        '井段End': calculation_End},
                                       index=[1])  # 自定义索引为：1 ，这里也可以不设置index
                df_temp.loc[:, "重计算厚度"] = df_temp.apply(self.get_thickness, axis=1)
                # 修改df_temp的最末一行
                df_temp.loc[len(df_temp), '井段(m)'] = ''.join([str(df_temp.loc[len(df_temp), '井段Start']),
                                                              '-', str(df_temp.loc[len(df_temp), '井段End'])])
                df_temp.loc[len(df_temp), '厚度(m)'] = df_temp.loc[len(df_temp), '重计算厚度']
            print(df_temp)
            ratio_Series = df_temp.groupby(by=['结论'])['重计算厚度'].sum() / df_temp['重计算厚度'].sum() * 100
            if ratio_Series.__len__() == 2:
                if '好' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                elif '中' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                elif '差' not in ratio_Series:
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
            elif ratio_Series.__len__() == 1:
                if ('好' not in ratio_Series) & ('中' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                elif ('好' not in ratio_Series) & ('差' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'好': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))
                elif ('中' not in ratio_Series) & ('差' not in ratio_Series):
                    ratio_Series = ratio_Series.append(pd.Series({'中': 0}))
                    ratio_Series = ratio_Series.append(pd.Series({'差': 0}))

        # 统计结论
        actual_Hao = str(round((calculation_End - calculation_Start) * (ratio_Series['好'] / 100), 2))
        Hao_Ratio = str(round(ratio_Series['好'], 2))

        actual_Zhong = str(round((calculation_End - calculation_Start) * (ratio_Series['中'] / 100), 2))
        Zhong_Ratio = str(round(ratio_Series['中'], 2))

        actual_Cha = str(round(calculation_End - calculation_Start - float(actual_Hao) - float(actual_Zhong), 2))
        Cha_Ratio = str(round(100.00 - float(Hao_Ratio) - float(Zhong_Ratio), 2))

        PATH = '.\\resources\\模板\\'
        wb = openpyxl.load_workbook(PATH + '2统模板.xlsx')
        sheet = wb[wb.sheetnames[0]]
        sheet['A1'] = ''.join(['二界面水泥胶结统计表（', str(calculation_Start), '-', str(calculation_End), 'm）'])
        sheet['C4'] = actual_Hao
        sheet['D4'] = Hao_Ratio
        sheet['C5'] = actual_Zhong
        sheet['D5'] = Zhong_Ratio
        sheet['C6'] = actual_Cha
        sheet['D6'] = Cha_Ratio

        self.mkdir('.\\WorkSpace\\合并统计工区')
        wb.save(
            '.\\WorkSpace\\合并统计工区\\二界面水泥胶结统计表(' + str(calculation_Start) + '-' + str(calculation_End) + 'm).xlsx')

        # 保存指定起始截止深度的单层统计表
        df_temp.drop(['井段Start', '井段End', '重计算厚度'], axis=1, inplace=True)
        df_temp.reset_index(drop=True, inplace=True)  # 重新设置列索引
        df_temp.index = df_temp.index + 1
        writer = pd.ExcelWriter(
            '.\\WorkSpace\\合并统计工区\\二界面水泥胶结单层评价表(' + str(calculation_Start) + '-' + str(calculation_End) + 'm).xlsx')
        df_temp.to_excel(writer, 'Sheet1')
        writer.save()

        # 单层统计表保存为Excel
        # df_all.drop(['井段Start', '井段End'], axis=1, inplace=True)
        # df_all.index = df_all.index + 1
        # writer = pd.ExcelWriter(
        #     '.\\WorkSpace\\合并统计工区\\单层评价表(合并)(' + str(start_Evaluation) + '-' + str(end_Evaluation) + 'm).xlsx')
        # df_all.to_excel(writer, 'Sheet1')
        # writer.save()

        QMessageBox.information(self, "提示", "运行完毕，请查看WorkSpace")

    def open_table_process_directory(self):
        path = '.\\WorkSpace\\合并统计工区'
        if not os.path.exists(path):
            os.makedirs(path)
            print(path, ' has been created.')
        os.startfile(path)

    def open_table_fusion_directory(self):
        path = '.\\WorkSpace\\综合评价工区'
        if not os.path.exists(path):
            os.makedirs(path)
            print(path, ' has been created.')
        os.startfile(path)

    def table_fusion_reaction(self):
        fileDir1 = self.lineEdit_75.text()
        fileDir2 = self.lineEdit_73.text()

        # 获取一界面单层评价表的深度界限
        df1 = pd.read_excel(fileDir1, header=2)
        df1.drop([0], inplace=True)
        df1 = df1.dropna(axis=0, how='any')  # 删除dataframe里NaN的所有行
        df1.loc[:, '井段(m)'] = df1['井段(m)'].str.replace(' ', '')  # 消除数据中空格
        # if len(df1) % 2 == 0:#如果len(df1)为偶数需要删除最后一行NaN，一行的情况不用删
        #     df1.drop([len(df1)], inplace=True)
        df1['井段Start'] = df1['井段(m)'].map(lambda x: x.split("-")[0])
        df1['井段End'] = df1['井段(m)'].map(lambda x: x.split("-")[1])
        # 表格数据清洗
        df1.loc[:, "井段Start"] = df1["井段Start"].str.replace(" ", "").astype('float')
        df1.loc[:, "井段End"] = df1["井段End"].str.replace(" ", "").astype('float')

        # 获取二界面单层评价表的深度界限
        df2 = pd.read_excel(fileDir2, header=2)
        df2.drop([0], inplace=True)
        # if df2.loc[1, '结论'] == '不确定':
        #     df2.drop([1], inplace=True)
        df2 = df2.dropna(axis=0, how='any')  # 删除dataframe里NaN的所有行
        df2.loc[:, '井段(m)'] = df2['井段(m)'].str.replace(' ', '')  # 消除数据中空格
        # if len(df2) % 2 == 0:#如果len(df2)为偶数需要删除最后一行NaN，一行的情况不用删
        #     df2.drop([len(df2)], inplace=True)
        df2['井段Start'] = df2['井段(m)'].map(lambda x: x.split("-")[0])
        df2['井段End'] = df2['井段(m)'].map(lambda x: x.split("-")[1])
        # 表格数据清洗
        df2.loc[:, "井段Start"] = df2["井段Start"].str.replace(" ", "").astype('float')
        df2.loc[:, "井段End"] = df2["井段End"].str.replace(" ", "").astype('float')

        list1 = df1['井段Start'].values.tolist()
        list2 = df1['井段End'].values.tolist()
        list3 = df2['井段Start'].values.tolist()
        list4 = df2['井段End'].values.tolist()
        # list合并并去重
        for item in list2:
            if item not in list1:
                list1.append(item)
        for item in list3:
            if item not in list1:
                list1.append(item)
        for item in list4:
            if item not in list1:
                list1.append(item)
        list1.sort(key=lambda x: float(x))
        print(list1)
        data = pd.DataFrame()
        for i in range(0, len(list1) - 1):
            j = i + 1
            evaluation_of_formation1 = self.layer_evaluation1(df1, list1[i], list1[j])[1]  # 调取一界面评价函数
            evaluation_of_formation2 = self.layer_evaluation2(df2, list1[i], list1[j])[1]  # 调取二界面评价函数
            if evaluation_of_formation1 == '好' and evaluation_of_formation2 in['好', '中', '不确定']:
                evaluation_of_formation = '优'
            elif evaluation_of_formation1 == '中' and evaluation_of_formation2 =='好':
                evaluation_of_formation = '优'
            elif evaluation_of_formation1 == '中' and evaluation_of_formation2 in ['中', '不确定']:
                evaluation_of_formation = '中等'
            elif evaluation_of_formation1 == '差' or evaluation_of_formation2 =='差':
                evaluation_of_formation = '差'
            thickness = round(list1[j] - list1[i], 2)
            interval = '-'.join([('%.2f' % list1[i]), ('%.2f' % list1[j])])
            print(interval, thickness, evaluation_of_formation1, evaluation_of_formation2, evaluation_of_formation, '\n')
            series = pd.Series({"井段(m)": interval, "厚度(m)": thickness, "一界面评价": evaluation_of_formation1, "二界面评价": evaluation_of_formation2, "综合评价": evaluation_of_formation}, name=i+1)
            data = data.append(series)
        # dataframe排序
        data = data[['井段(m)', '厚度(m)', '一界面评价', '二界面评价', '综合评价']]
        print(data)

        # 保存为excel
        writer = pd.ExcelWriter('.\\WorkSpace\\综合评价工区\\综合评价表.xlsx')
        data.to_excel(writer, 'Sheet1')
        writer.save()
        QMessageBox.information(self, "提示", "运行完毕，请查看工区")

    def search_for_statistic_result(self):
        start_depth = float(self.lineEdit_78.text())
        end_depth = float(self.lineEdit_79.text())

        if self.lineEdit_77.text() != '' and self.lineEdit_76.text() != '':
            QMessageBox.information(self, "提示", "暂不支持两个评价表同时统计，请删除一个后重试")
        elif self.lineEdit_77.text() != '':
            fileDir1 = self.lineEdit_77.text()

            # 获取一界面单层评价表的深度界限
            df1 = pd.read_excel(fileDir1, header=2)
            df1.drop([0], inplace=True)
            df1 = df1.dropna(axis=0, how='any')  # 删除dataframe里NaN的所有行
            df1.loc[:, '井段(m)'] = df1['井段(m)'].str.replace(' ', '')  # 消除数据中空格
            # if len(df1) % 2 == 0:#如果len(df1)为偶数需要删除最后一行NaN，一行的情况不用删
            #     df1.drop([len(df1)], inplace=True)
            df1['井段Start'] = df1['井段(m)'].map(lambda x: x.split("-")[0])
            df1['井段End'] = df1['井段(m)'].map(lambda x: x.split("-")[1])
            # 表格数据清洗
            df1.loc[:, "井段Start"] = df1["井段Start"].str.replace(" ", "").astype('float')
            df1.loc[:, "井段End"] = df1["井段End"].str.replace(" ", "").astype('float')
            evaluation_of_formation1 = self.layer_evaluation1(df1, start_depth, end_depth)[0]  # 调取一界面评价函数
            self.lineEdit_83.setText(('%.2f' % evaluation_of_formation1['好']))
            self.lineEdit_81.setText(('%.2f' % evaluation_of_formation1['中']))
            self.lineEdit_80.setText(('%.2f' % evaluation_of_formation1['差']))
            not_sure = 100 - evaluation_of_formation1['好'] - evaluation_of_formation1['中'] - evaluation_of_formation1['差']
            self.lineEdit_82.setText(('%.2f' % not_sure))

        elif self.lineEdit_76.text() != '':
            fileDir2 = self.lineEdit_76.text()

            # 获取一界面单层评价表的深度界限
            df1 = pd.read_excel(fileDir2, header=2)
            df1.drop([0], inplace=True)
            df1 = df1.dropna(axis=0, how='any')  # 删除dataframe里NaN的所有行
            df1.loc[:, '井段(m)'] = df1['井段(m)'].str.replace(' ', '')  # 消除数据中空格
            # if len(df1) % 2 == 0:#如果len(df1)为偶数需要删除最后一行NaN，一行的情况不用删
            #     df1.drop([len(df1)], inplace=True)
            df1['井段Start'] = df1['井段(m)'].map(lambda x: x.split("-")[0])
            df1['井段End'] = df1['井段(m)'].map(lambda x: x.split("-")[1])
            # 表格数据清洗
            df1.loc[:, "井段Start"] = df1["井段Start"].str.replace(" ", "").astype('float')
            df1.loc[:, "井段End"] = df1["井段End"].str.replace(" ", "").astype('float')
            evaluation_of_formation2 = self.layer_evaluation2(df1, start_depth, end_depth)[0]  # 调取二界面评价函数
            self.lineEdit_83.setText(('%.2f' % evaluation_of_formation2['好']))
            self.lineEdit_81.setText(('%.2f' % evaluation_of_formation2['中']))
            self.lineEdit_80.setText(('%.2f' % evaluation_of_formation2['差']))
            not_sure = 100 - evaluation_of_formation2['好'] - evaluation_of_formation2['中'] - evaluation_of_formation2['差']
            self.lineEdit_82.setText(('%.2f' % not_sure))

if __name__ == "__main__":
    app = QApplication(sys.argv)
    main = Main_window()
    main.show()
    sys.exit(app.exec_())